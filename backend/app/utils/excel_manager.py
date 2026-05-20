import openpyxl
import os
from sqlalchemy.orm import Session
from app.models.producto import Producto
from app.models.listas_precios import ListaPreciosDetalle, ListaPrecios
import datetime

def import_prices_from_excel(db: Session, file_path: str, lista_id: int) -> int:
    """
    Imports products and price details from an Excel file.
    Columns expected: Codigo, Descripcion, Precio Costo, Precio Venta, Precio Mayoreo, Existencia, Inv. Minimo, Departamento
    """
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active
    
    # Identify headers mapping
    headers = [cell.value for cell in sheet[1]]
    
    col_mapping = {}
    expected_cols = {
        "codigo": ["codigo", "code", "cod"],
        "descripcion": ["descripcion", "descripción", "description", "desc"],
        "precio_costo": ["precio costo", "costo", "precio_costo", "cost"],
        "precio_venta": ["precio venta", "venta", "precio_venta", "price"],
        "precio_mayoreo": ["precio mayoreo", "mayoreo", "precio_mayoreo", "wholesale"],
        "existencia": ["existencia", "stock", "cant", "cantidad"],
        "inv_minimo": ["inv. minimo", "inv_minimo", "stock_minimo", "minimo", "mínimo"],
        "departamento": ["departamento", "depto", "categoría", "categoria", "family", "familia"]
    }
    
    for key, aliases in expected_cols.items():
        col_mapping[key] = None
        for i, header in enumerate(headers):
            if header and str(header).lower().strip() in aliases:
                col_mapping[key] = i + 1
                break
                
    # Fallback to column index if headers are not exact
    if col_mapping["codigo"] is None: col_mapping["codigo"] = 1
    if col_mapping["descripcion"] is None: col_mapping["descripcion"] = 2
    if col_mapping["precio_costo"] is None: col_mapping["precio_costo"] = 3
    if col_mapping["precio_venta"] is None: col_mapping["precio_venta"] = 4
    if col_mapping["precio_mayoreo"] is None: col_mapping["precio_mayoreo"] = 5
    if col_mapping["existencia"] is None: col_mapping["existencia"] = 6
    if col_mapping["inv_minimo"] is None: col_mapping["inv_minimo"] = 7
    if col_mapping["departamento"] is None: col_mapping["departamento"] = 8

    imported_count = 0
    
    # Process rows (skipping header)
    for r in range(2, sheet.max_row + 1):
        raw_code = sheet.cell(row=r, column=col_mapping["codigo"]).value
        if raw_code is None:
            continue
            
        code_str = str(raw_code).strip()
        
        # Clean prefix if it matches list prefix (e.g. List 2 starts with '2', so '2001' becomes '001')
        base_code = code_str
        prefix_str = str(lista_id)
        if code_str.startswith(prefix_str) and len(code_str) > len(prefix_str):
            base_code = code_str[len(prefix_str):]
            
        desc = sheet.cell(row=r, column=col_mapping["descripcion"]).value
        if desc is None:
            desc = f"Producto {base_code}"
        else:
            desc = str(desc).strip()
            
        depto = sheet.cell(row=r, column=col_mapping["departamento"]).value
        depto = str(depto).strip() if depto else "Cortes frescos"
        
        # Parse numeric values helper
        def parse_float(val, default=0.0):
            if val is None:
                return default
            try:
                return float(str(val).replace(",", ".").replace("$", "").strip())
            except ValueError:
                return default

        costo = parse_float(sheet.cell(row=r, column=col_mapping["precio_costo"]).value)
        venta = parse_float(sheet.cell(row=r, column=col_mapping["precio_venta"]).value)
        mayoreo = parse_float(sheet.cell(row=r, column=col_mapping["precio_mayoreo"]).value)
        stock = parse_float(sheet.cell(row=r, column=col_mapping["existencia"]).value)
        min_stock = parse_float(sheet.cell(row=r, column=col_mapping["inv_minimo"]).value)

        # 1. Lookup or create Base Product
        prod = db.query(Producto).filter(Producto.codigo == base_code).first()
        if not prod:
            prod = Producto(
                codigo=base_code,
                descripcion=desc,
                departamento=depto,
                activo=True
            )
            db.add(prod)
            db.commit()
            db.refresh(prod)
        else:
            # Update desc & family
            prod.descripcion = desc
            prod.departamento = depto
            db.commit()

        # 2. Lookup or create prices details for this list
        det = db.query(ListaPreciosDetalle).filter(
            ListaPreciosDetalle.lista_precios_id == lista_id,
            ListaPreciosDetalle.producto_id == prod.id
        ).first()
        
        if not det:
            det = ListaPreciosDetalle(
                lista_precios_id=lista_id,
                producto_id=prod.id,
                precio_costo=costo,
                precio_venta=venta,
                precio_mayoreo=mayoreo,
                stock=stock,
                stock_minimo=min_stock
            )
            db.add(det)
        else:
            det.precio_costo = costo
            det.precio_venta = venta
            det.precio_mayoreo = mayoreo
            det.stock = stock
            det.stock_minimo = min_stock
            
        imported_count += 1
        
    # Update pricing list date
    lista_obj = db.query(ListaPrecios).filter(ListaPrecios.id == lista_id).first()
    if lista_obj:
        lista_obj.fecha_actualizacion = datetime.datetime.utcnow()
        
    db.commit()
    return imported_count

def export_prices_to_excel(db: Session, lista_id: int, upload_dir: str) -> str:
    """
    Exports a pricing list with all product details to an Excel file.
    """
    lista = db.query(ListaPrecios).filter(ListaPrecios.id == lista_id).first()
    if not lista:
        raise ValueError("Lista de precios no encontrada")
        
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = f"Lista_{lista_id}"
    
    # Set headers
    headers = ["Codigo", "Descripcion", "Precio Costo", "Precio Venta", "Precio Mayoreo", "Existencia", "Inv. Minimo", "Departamento"]
    sheet.append(headers)
    
    # Query details
    detalles = db.query(ListaPreciosDetalle).filter(ListaPreciosDetalle.lista_precios_id == lista_id).all()
    
    for det in detalles:
        # Prefix the product code
        prefixed_code = f"{lista_id}{det.producto.codigo}"
        sheet.append([
            prefixed_code,
            det.producto.descripcion,
            det.precio_costo,
            det.precio_venta,
            det.precio_mayoreo,
            det.stock,
            det.stock_minimo,
            det.producto.departamento
        ])
        
    filename = f"Exportacion_Lista_{lista_id}_{datetime.date.today().strftime('%Y%m%d')}.xlsx"
    dest_path = os.path.join(upload_dir, filename)
    wb.save(dest_path)
    
    return f"/uploads/{filename}"
